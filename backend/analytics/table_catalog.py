"""Persistent table asset catalog for the analytics workbench.

The upload/import entry remains the existing knowledge import flow. This module
keeps spreadsheet-like assets in the catalog database so the analytics page and
Pandas tool do not repeatedly scan `/knowledge`.
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from knowledge.models import KnowledgeDocument, KnowledgeTableAsset
from knowledge.paths import get_knowledge_root
from knowledge.service import DEFAULT_KNOWLEDGE_BASE_ID, KnowledgeService
from utils.table_engine.profiler import profile_dataframe

TABLE_SUFFIXES = {".xlsx", ".xls", ".csv", ".tsv"}
DERIVED_CONCAT_SOURCE_TYPE = "derived_concat"
LOGICAL_CONCAT_SOURCE_TYPE = "logical_concat"
CONCAT_SOURCE_TYPES = {DERIVED_CONCAT_SOURCE_TYPE, LOGICAL_CONCAT_SOURCE_TYPE}
DERIVED_CONCAT_LINEAGE_COLUMNS = (
    "_pc_source_asset_id",
    "_pc_source_file_name",
    "_pc_source_sheet_name",
    "_pc_source_virtual_path",
)
IGNORED_TOP_LEVEL_DIRS = {".puddingclaw", ".tasks", "tasks", "originals", "assets"}
PROFILE_SAMPLE_ROWS = 20000
BEIJING_TZ = ZoneInfo("Asia/Shanghai")


def _utc_datetime(timestamp: float) -> datetime:
    return datetime.fromtimestamp(timestamp, tz=timezone.utc)


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _virtual_path(root: Path, path: Path) -> str:
    return f"/knowledge/{path.relative_to(root).as_posix()}"


def _profile_dir(root: Path) -> Path:
    return root / ".puddingclaw" / "table_profiles"


def _asset_id(virtual_path: str, sheet_name: str | None) -> str:
    payload = f"{virtual_path}#{sheet_name or ''}"
    return "tbl_" + hashlib.sha1(payload.encode("utf-8")).hexdigest()[:24]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_default(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


@dataclass(frozen=True)
class TableAssetRef:
    asset_id: str
    path: Path
    virtual_path: str
    file_name: str
    source_type: str
    sheet_name: str | None
    size_bytes: int
    modified_at: datetime
    content_sha256: str = ""
    document_id: str | None = None


class TableCatalogError(RuntimeError):
    """Raised when table catalog operations fail."""


class TableAssetCatalog:
    """Catalog imported table files and generate cached profiles."""

    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        self.knowledge_root = get_knowledge_root(base_dir).expanduser().resolve()

    def profile_path(self, asset_id: str) -> Path:
        return _profile_dir(self.knowledge_root) / f"{asset_id}.profile.json"

    async def list_assets(
        self,
        session: AsyncSession,
        *,
        include_profile: bool = False,
        limit: int = 500,
        ensure_scanned: bool = True,
    ) -> list[dict[str, Any]]:
        if ensure_scanned:
            await self.ensure_catalog_populated(session, limit=limit)
        stmt = (
            select(KnowledgeTableAsset)
            .where(
                KnowledgeTableAsset.knowledge_base_id == DEFAULT_KNOWLEDGE_BASE_ID,
                KnowledgeTableAsset.reference_status != "removed",
            )
            .order_by(KnowledgeTableAsset.updated_at.desc(), KnowledgeTableAsset.file_name.asc())
            .limit(max(1, min(limit, 2000)))
        )
        result = await session.execute(stmt)
        return [self._asset_to_dict(asset, include_profile=include_profile) for asset in result.scalars()]

    async def get_asset(self, session: AsyncSession, asset_id: str, *, include_profile: bool = True) -> dict[str, Any]:
        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None:
            await self.ensure_catalog_populated(session, limit=2000)
            asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed":
            raise TableCatalogError("Table asset not found.")
        return self._asset_to_dict(asset, include_profile=include_profile)

    async def remove_asset(self, session: AsyncSession, asset_id: str) -> dict[str, Any]:
        """Detach one uploaded workbook/flat file from analytics without deleting its KB file."""
        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed":
            raise TableCatalogError("Table asset not found.")

        siblings = list(
            (
                await session.execute(
                    select(KnowledgeTableAsset).where(
                        KnowledgeTableAsset.knowledge_base_id == asset.knowledge_base_id,
                        KnowledgeTableAsset.storage_path == asset.storage_path,
                        KnowledgeTableAsset.reference_status != "removed",
                    )
                )
            ).scalars()
        )
        removed_ids: list[str] = []
        for sibling in siblings:
            profile_path = Path(sibling.profile_path) if sibling.profile_path else self.profile_path(sibling.asset_id)
            if profile_path.is_file() and self.profile_path(sibling.asset_id).parent in profile_path.parents:
                await asyncio.to_thread(profile_path.unlink, missing_ok=True)
            sibling.reference_status = "removed"
            sibling.profile_status = "missing"
            sibling.profile_path = ""
            sibling.rows = None
            sibling.columns_count = None
            sibling.columns = []
            removed_ids.append(sibling.asset_id)
        await session.commit()
        return {
            "asset_id": asset_id,
            "removed_asset_ids": removed_ids,
            "file_name": asset.file_name,
            "storage_path": asset.storage_path,
            "source_file_preserved": True,
        }

    async def load_dataframe_for_asset(self, session: AsyncSession, asset_id: str):
        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None:
            await self.ensure_catalog_populated(session, limit=2000)
            asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed":
            raise TableCatalogError("Table asset not found.")
        if asset.source_type == LOGICAL_CONCAT_SOURCE_TYPE:
            logical = (asset.asset_metadata or {}).get("logical_dataset") if isinstance(asset.asset_metadata, dict) else None
            source_ids = logical.get("source_asset_ids") if isinstance(logical, dict) else []
            if not isinstance(source_ids, list) or len(source_ids) < 2:
                raise TableCatalogError("虚拟逻辑数据集缺少有效来源定义。")
            result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
            by_id = {item.asset_id: item for item in result.scalars()}
            if any(source_id not in by_id or by_id[source_id].reference_status == "removed" for source_id in source_ids):
                raise TableCatalogError("虚拟逻辑数据集存在不可用来源表。")
            refs = [self._ref_from_model(by_id[source_id]) for source_id in source_ids]
            return asset, await asyncio.to_thread(
                self._read_virtual_concat,
                refs,
                schema_mode=str(logical.get("schema_mode") or "strict"),
                canonical_columns=[str(item) for item in logical.get("canonical_columns") or []],
            )
        return asset, self._load_dataframe(self._ref_from_model(asset))

    async def create_concat_dataset(
        self,
        session: AsyncSession,
        *,
        name: str,
        description: str = "",
        tags: list[str] | None = None,
        source_asset_ids: list[str],
        schema_mode: str = "strict",
        routing: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Materialize periodic table assets as one vertical logical dataset."""

        clean_name = name.strip()
        source_ids = list(dict.fromkeys(item.strip() for item in source_asset_ids if item and item.strip()))
        if not clean_name:
            raise TableCatalogError("逻辑数据集名称不能为空。")
        if len(source_ids) < 2:
            raise TableCatalogError("垂直合并至少需要选择两张表格资产。")
        if schema_mode not in {"strict", "baseline_fill_missing", "union_fill_missing"}:
            raise TableCatalogError("不支持的字段合并方式。")

        result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
        by_id = {item.asset_id: item for item in result.scalars()}
        missing = [asset_id for asset_id in source_ids if asset_id not in by_id or by_id[asset_id].reference_status == "removed"]
        if missing:
            raise TableCatalogError(f"来源表不存在或已移除: {', '.join(missing)}")
        sources = [by_id[asset_id] for asset_id in source_ids]
        if any(source.source_type in CONCAT_SOURCE_TYPES for source in sources):
            raise TableCatalogError("第一版不支持嵌套合并数据集；请直接选择原始表格来源。")

        schema_preview = await asyncio.to_thread(self._inspect_concat_refs, [self._ref_from_model(source) for source in sources])
        if schema_preview["has_schema_drift"] and schema_mode == "strict":
            raise TableCatalogError(
                "所选表字段不完全一致。请先确认“缺失字段补空并合并”，"
                "缺失字段会置空，额外字段会保留在逻辑数据集内。"
            )

        dataset_id = "tbl_concat_" + hashlib.sha1(clean_name.encode("utf-8")).hexdigest()[:24]
        existing = await session.get(KnowledgeTableAsset, dataset_id)
        if existing and existing.reference_status != "removed":
            raise TableCatalogError("同名逻辑数据集已存在；请使用刷新操作或更换名称。")

        refs = [self._ref_from_model(source) for source in sources]
        from runtime_identity.paths import PuddingClawPaths

        output_dir = PuddingClawPaths.from_environment().data() / "analytics-concat-datasets" / dataset_id
        output_path = output_dir / "dataset.json"
        output_columns = (
            schema_preview["baseline_columns"]
            if schema_mode == "baseline_fill_missing"
            else schema_preview["canonical_columns"]
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        total_rows = sum(int(source.rows or 0) for source in sources) if all(source.rows is not None for source in sources) else None

        now = datetime.now(BEIJING_TZ)
        logical_definition = self._logical_dataset_definition(
            asset_id=dataset_id,
            name=clean_name,
            description=description,
            tags=tags or [],
            routing=routing,
            sources=sources,
            source_ids=source_ids,
            schema_mode=schema_mode,
            output_columns=output_columns,
            schema_preview=schema_preview,
            refreshed_at=now,
        )
        metadata = {"logical_dataset": logical_definition}
        definition = {"asset_id": dataset_id, "name": clean_name, **logical_definition}
        await asyncio.to_thread(output_path.write_text, json.dumps(definition, ensure_ascii=False, indent=2), encoding="utf-8")
        if existing is None:
            existing = KnowledgeTableAsset(asset_id=dataset_id, knowledge_base_id=DEFAULT_KNOWLEDGE_BASE_ID)
            session.add(existing)
        existing.document_id = None
        existing.source_type = LOGICAL_CONCAT_SOURCE_TYPE
        existing.file_name = clean_name
        existing.storage_path = str(output_path)
        existing.virtual_path = f"/knowledge/.puddingclaw/derived/concat/{dataset_id}/dataset.json"
        existing.sheet_name = None
        existing.size_bytes = output_path.stat().st_size
        existing.modified_at = now
        existing.content_sha256 = await asyncio.to_thread(_sha256, output_path)
        existing.profile_status = "missing"
        existing.profile_path = ""
        existing.rows = total_rows
        existing.columns_count = len(output_columns) + len(DERIVED_CONCAT_LINEAGE_COLUMNS)
        existing.columns = [*output_columns, *DERIVED_CONCAT_LINEAGE_COLUMNS]
        existing.reference_status = "ready"
        existing.asset_metadata = metadata
        await session.commit()
        await session.refresh(existing)
        return self._asset_to_dict(existing, include_profile=True)

    async def preview_concat_dataset(
        self,
        session: AsyncSession,
        *,
        source_asset_ids: list[str],
    ) -> dict[str, Any]:
        """Inspect selected inputs before the user confirms a non-strict concat."""

        source_ids = list(dict.fromkeys(item.strip() for item in source_asset_ids if item and item.strip()))
        if len(source_ids) < 2:
            raise TableCatalogError("垂直合并至少需要选择两张表格资产。")
        result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
        by_id = {item.asset_id: item for item in result.scalars()}
        missing = [asset_id for asset_id in source_ids if asset_id not in by_id or by_id[asset_id].reference_status == "removed"]
        if missing:
            raise TableCatalogError(f"来源表不存在或已移除: {', '.join(missing)}")
        sources = [by_id[asset_id] for asset_id in source_ids]
        if any(source.source_type in CONCAT_SOURCE_TYPES for source in sources):
            raise TableCatalogError("第一版不支持嵌套合并数据集；请直接选择原始表格来源。")
        return await asyncio.to_thread(self._inspect_concat_refs, [self._ref_from_model(source) for source in sources])

    async def refresh_concat_dataset(
        self,
        session: AsyncSession,
        asset_id: str,
        *,
        source_asset_ids: list[str] | None = None,
        schema_mode_override: str | None = None,
    ) -> dict[str, Any]:
        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed" or asset.source_type not in CONCAT_SOURCE_TYPES:
            raise TableCatalogError("逻辑合并数据集不存在。")
        metadata = dict(asset.asset_metadata) if isinstance(asset.asset_metadata, dict) else {}
        logical = dict(metadata.get("logical_dataset")) if isinstance(metadata.get("logical_dataset"), dict) else {}
        source_ids = source_asset_ids if source_asset_ids is not None else logical.get("source_asset_ids")
        source_ids = list(dict.fromkeys(str(item).strip() for item in source_ids if str(item).strip())) if isinstance(source_ids, list) else []
        if len(source_ids) < 2:
            raise TableCatalogError("逻辑数据集缺少有效来源定义。")
        # Recreate in-place from the registered sources while preserving its asset ID.
        result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
        by_id = {item.asset_id: item for item in result.scalars()}
        if any(asset_id not in by_id or by_id[asset_id].reference_status == "removed" for asset_id in source_ids):
            raise TableCatalogError("逻辑数据集存在已移除的来源表，请先修复来源后刷新。")
        refs = [self._ref_from_model(by_id[source_id]) for source_id in source_ids]
        if any(ref.source_type in CONCAT_SOURCE_TYPES for ref in refs):
            raise TableCatalogError("不支持把逻辑数据集继续嵌套合并。")
        schema_mode = schema_mode_override or str(logical.get("schema_mode") or "strict")
        if schema_mode not in {"strict", "baseline_fill_missing", "union_fill_missing"}:
            raise TableCatalogError("不支持的字段合并方式。")
        schema_preview = await asyncio.to_thread(self._inspect_concat_refs, refs)
        if schema_preview["has_schema_drift"] and schema_mode == "strict":
            raise TableCatalogError("来源字段发生变化，请重新确认“缺失字段补空并合并”后新建逻辑数据集。")
        output_columns = (
            schema_preview["baseline_columns"]
            if schema_mode == "baseline_fill_missing"
            else schema_preview["canonical_columns"]
        )
        if asset.source_type == LOGICAL_CONCAT_SOURCE_TYPE:
            output_path = Path(asset.storage_path)
            total_rows = sum(int(by_id[source_id].rows or 0) for source_id in source_ids) if all(by_id[source_id].rows is not None for source_id in source_ids) else None
            now = datetime.now(BEIJING_TZ)
            logical = self._logical_dataset_definition(
                asset_id=asset.asset_id,
                name=asset.file_name,
                description=str(logical.get("description") or ""),
                tags=[str(item) for item in logical.get("tags") or []],
                routing=logical.get("routing") if isinstance(logical.get("routing"), dict) else None,
                sources=[by_id[source_id] for source_id in source_ids],
                source_ids=source_ids,
                schema_mode=schema_mode,
                output_columns=output_columns,
                schema_preview=schema_preview,
                refreshed_at=now,
            )
            definition = {"asset_id": asset.asset_id, "name": asset.file_name, **logical}
            await asyncio.to_thread(output_path.write_text, json.dumps(definition, ensure_ascii=False, indent=2), encoding="utf-8")
            asset.asset_metadata = {**metadata, "logical_dataset": logical}
            asset.size_bytes = output_path.stat().st_size
            asset.modified_at = now
            asset.content_sha256 = await asyncio.to_thread(_sha256, output_path)
            asset.rows = total_rows
            asset.columns_count = len(output_columns) + len(DERIVED_CONCAT_LINEAGE_COLUMNS)
            asset.columns = [*output_columns, *DERIVED_CONCAT_LINEAGE_COLUMNS]
            await session.commit()
            await session.refresh(asset)
            return self._asset_to_dict(asset, include_profile=True)

        output_path = Path(asset.storage_path)
        materialized = await asyncio.to_thread(
            self._materialize_concat_dataset,
            refs,
            output_path,
            schema_mode=schema_mode,
            canonical_columns=output_columns,
        )
        profile = await asyncio.to_thread(self._build_profile, None, self._derived_ref(asset.asset_id, asset.file_name, output_path, materialized["rows"]))
        profile_path = self.profile_path(asset.asset_id)
        await asyncio.to_thread(profile_path.write_text, json.dumps(profile, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")
        now = datetime.now(BEIJING_TZ)
        logical.update({
            "schema_mode": schema_mode,
            "source_asset_ids": source_ids,
            "canonical_columns": output_columns,
            "canonical_columns": materialized["canonical_columns"],
            "schema_preview": schema_preview,
            "refreshed_at": now.isoformat(),
        })
        asset.asset_metadata = {**metadata, "logical_dataset": logical}
        asset.size_bytes = output_path.stat().st_size
        asset.modified_at = now
        asset.content_sha256 = await asyncio.to_thread(_sha256, output_path)
        asset.profile_status = "ready"
        asset.profile_path = str(profile_path)
        asset.rows = materialized["rows"]
        asset.columns_count = len(materialized["columns"])
        asset.columns = materialized["columns"]
        await session.commit()
        await session.refresh(asset)
        return self._asset_to_dict(asset, include_profile=True)

    async def append_concat_dataset_sources(
        self,
        session: AsyncSession,
        *,
        asset_id: str,
        source_asset_ids: list[str],
        schema_mode: str,
    ) -> dict[str, Any]:
        """Add later period tables to an existing logical dataset and rematerialize it."""

        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed" or asset.source_type not in CONCAT_SOURCE_TYPES:
            raise TableCatalogError("逻辑合并数据集不存在。")
        metadata = dict(asset.asset_metadata) if isinstance(asset.asset_metadata, dict) else {}
        logical = dict(metadata.get("logical_dataset")) if isinstance(metadata.get("logical_dataset"), dict) else {}
        existing_ids = logical.get("source_asset_ids") if isinstance(logical.get("source_asset_ids"), list) else []
        append_ids = list(dict.fromkeys(item.strip() for item in source_asset_ids if item and item.strip()))
        if not append_ids:
            raise TableCatalogError("请至少选择一张要追加的来源表。")
        new_ids = [item for item in append_ids if item not in existing_ids]
        if not new_ids:
            raise TableCatalogError("所选来源已在此逻辑数据集中。")
        return await self.refresh_concat_dataset(
            session,
            asset_id,
            source_asset_ids=[*existing_ids, *new_ids],
            schema_mode_override=schema_mode,
        )

    async def update_logical_dataset_definition(
        self,
        session: AsyncSession,
        *,
        asset_id: str,
        name: str | None = None,
        description: str | None = None,
        tags: list[str] | None = None,
        preferred_intents: list[str] | None = None,
        direct_source_allowed: bool | None = None,
    ) -> dict[str, Any]:
        """Update business metadata only; source bindings and data stay untouched."""

        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed" or asset.source_type != LOGICAL_CONCAT_SOURCE_TYPE:
            raise TableCatalogError("虚拟逻辑数据集不存在。")
        metadata = dict(asset.asset_metadata) if isinstance(asset.asset_metadata, dict) else {}
        logical = dict(metadata.get("logical_dataset")) if isinstance(metadata.get("logical_dataset"), dict) else {}
        source_ids = [str(item).strip() for item in logical.get("source_asset_ids") or [] if str(item).strip()]
        if len(source_ids) < 2:
            raise TableCatalogError("虚拟逻辑数据集缺少有效来源定义。")
        result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
        by_id = {item.asset_id: item for item in result.scalars()}
        if any(source_id not in by_id or by_id[source_id].reference_status == "removed" for source_id in source_ids):
            raise TableCatalogError("虚拟逻辑数据集存在不可用来源表。")

        next_name = (name if name is not None else asset.file_name).strip()
        if not next_name:
            raise TableCatalogError("逻辑数据集名称不能为空。")
        next_description = str(description if description is not None else logical.get("description") or "").strip()
        next_tags = [str(item).strip() for item in (tags if tags is not None else logical.get("tags") or []) if str(item).strip()]
        old_routing = logical.get("routing") if isinstance(logical.get("routing"), dict) else {}
        routing = {
            "preferred_intents": [str(item).strip() for item in (preferred_intents if preferred_intents is not None else old_routing.get("preferred_intents") or []) if str(item).strip()],
            "direct_source_allowed": bool(old_routing.get("direct_source_allowed", True) if direct_source_allowed is None else direct_source_allowed),
        }
        schema_preview = logical.get("schema_preview") if isinstance(logical.get("schema_preview"), dict) else {}
        output_columns = [str(item) for item in logical.get("canonical_columns") or (logical.get("schema") or {}).get("fields") or []]
        now = datetime.now(BEIJING_TZ)
        definition = self._logical_dataset_definition(
            asset_id=asset.asset_id,
            name=next_name,
            description=next_description,
            tags=next_tags,
            routing=routing,
            sources=[by_id[source_id] for source_id in source_ids],
            source_ids=source_ids,
            schema_mode=str(logical.get("schema_mode") or "strict"),
            output_columns=output_columns,
            schema_preview=schema_preview,
            refreshed_at=now,
        )
        # Keep the actual profile timestamp when metadata-only editing does not change sources.
        if isinstance(logical.get("profile"), dict):
            definition["profile"] = logical["profile"]
            definition["coverage"] = logical.get("coverage") or []
            if logical.get("profile_refreshed_at"):
                definition["profile_refreshed_at"] = logical["profile_refreshed_at"]
        definition_path = Path(asset.storage_path)
        await asyncio.to_thread(definition_path.write_text, json.dumps({"asset_id": asset.asset_id, "name": next_name, **definition}, ensure_ascii=False, indent=2), encoding="utf-8")
        asset.file_name = next_name
        asset.asset_metadata = {**metadata, "logical_dataset": definition}
        asset.size_bytes = definition_path.stat().st_size
        asset.modified_at = now
        asset.content_sha256 = await asyncio.to_thread(_sha256, definition_path)
        await session.commit()
        await session.refresh(asset)
        return self._asset_to_dict(asset, include_profile=True)

    async def generate_profile(
        self,
        session: AsyncSession,
        asset_id: str,
        *,
        include_profile: bool = True,
        populate_logical_source_profiles: bool = True,
    ) -> dict[str, Any]:
        asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None:
            await self.ensure_catalog_populated(session, limit=2000)
            asset = await session.get(KnowledgeTableAsset, asset_id)
        if asset is None or asset.reference_status == "removed":
            raise TableCatalogError("Table asset not found.")

        if asset.source_type == LOGICAL_CONCAT_SOURCE_TYPE:
            return await self._generate_logical_dataset_profile(
                session,
                asset,
                include_profile=include_profile,
                populate_missing_source_profiles=populate_logical_source_profiles,
            )

        ref = self._ref_from_model(asset)
        profile = await asyncio.to_thread(self._build_profile, asset, ref)
        profile_path = self.profile_path(asset.asset_id)
        profile_path.parent.mkdir(parents=True, exist_ok=True)
        profile_json = json.dumps(profile, ensure_ascii=False, indent=2, default=_json_default)
        await asyncio.to_thread(profile_path.write_text, profile_json, encoding="utf-8")

        shape = profile.get("shape") if isinstance(profile.get("shape"), list) else []
        columns = [column.get("name") for column in profile.get("columns", []) if isinstance(column, dict)]
        asset.profile_status = "ready"
        asset.profile_path = str(profile_path)
        asset.rows = int(shape[0]) if len(shape) > 0 and shape[0] is not None else None
        asset.columns_count = int(shape[1]) if len(shape) > 1 and shape[1] is not None else None
        asset.columns = [str(column) for column in columns if column]
        asset.asset_metadata = {
            **(asset.asset_metadata or {}),
            "profile_generated_at": profile["generated_at"],
        }
        await session.commit()
        await session.refresh(asset)
        return self._asset_to_dict(asset, include_profile=include_profile)

    def _build_profile(self, _asset: KnowledgeTableAsset | None, ref: TableAssetRef) -> dict[str, Any]:
        """Build a sampled table profile off the event loop."""

        df = self._load_dataframe(ref, sample_rows=PROFILE_SAMPLE_ROWS)
        base_profile = profile_dataframe(df, preview_rows=8)
        actual_shape = self._table_shape(ref, sample_df=df)
        sampled_rows = int(df.shape[0])
        total_rows = int(actual_shape[0]) if actual_shape and actual_shape[0] is not None else sampled_rows
        return {
            "asset_id": ref.asset_id,
            "kind": "table_asset_profile",
            "sampled": sampled_rows < total_rows,
            "sample_rows": sampled_rows,
            "profile_sample_limit": PROFILE_SAMPLE_ROWS,
            "source_type": ref.source_type,
            "file_name": ref.file_name,
            "virtual_path": ref.virtual_path,
            "sheet_name": ref.sheet_name,
            "size_bytes": ref.size_bytes,
            "modified_at": _iso(ref.modified_at),
            "generated_at": datetime.now(BEIJING_TZ).isoformat(),
            "source_file_state": self._source_file_state(ref.path),
            "source_content_sha256": _asset.content_sha256 if _asset is not None else ref.content_sha256,
            "shape": actual_shape or base_profile.get("shape"),
            "columns": self._column_profiles(df, base_profile),
            "dtypes": base_profile.get("dtypes", {}),
            "preview": base_profile.get("preview", []),
        }

    @staticmethod
    def _source_file_state(path: Path) -> dict[str, Any]:
        """Return a cheap freshness marker without rescanning table contents."""

        try:
            stat = path.stat()
        except OSError:
            return {"exists": False}
        return {
            "exists": True,
            "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=BEIJING_TZ).isoformat(),
            "size_bytes": int(stat.st_size),
        }

    @staticmethod
    def _is_temporal_field(name: str) -> bool:
        normalized = name.lower()
        return bool(re.search(r"日期|时间|年月|年周|月份|date|time|month|week|year", normalized))

    @staticmethod
    def _profile_source_state(source: KnowledgeTableAsset) -> tuple[dict[str, Any] | None, dict[str, Any]]:
        profile_path = Path(source.profile_path) if source.profile_path else None
        profile = TableAssetCatalog._read_profile(profile_path) if profile_path and profile_path.exists() else None
        current_state = TableAssetCatalog._source_file_state(Path(source.storage_path))
        profiled_state = profile.get("source_file_state") if isinstance(profile, dict) else None
        fresh = bool(profile)
        if profile and current_state.get("exists"):
            fresh = isinstance(profiled_state, dict) and (
                profiled_state.get("modified_at") == current_state.get("modified_at")
                and profiled_state.get("size_bytes") == current_state.get("size_bytes")
            )
        return profile, {
            "asset_id": source.asset_id,
            "name": source.file_name,
            "sheet_name": source.sheet_name,
            "profile_status": "fresh" if profile and fresh else ("stale" if profile else "missing"),
            "profile_generated_at": profile.get("generated_at") if isinstance(profile, dict) else None,
            "source_file_state": current_state,
        }

    @classmethod
    def _profile_coverage_for_source(cls, profile: dict[str, Any] | None) -> list[dict[str, Any]]:
        if not isinstance(profile, dict):
            return []
        dimensions: list[dict[str, Any]] = []
        for column in profile.get("columns") or []:
            if not isinstance(column, dict):
                continue
            name = str(column.get("name") or "").strip()
            observed_range = column.get("observed_range") if isinstance(column.get("observed_range"), dict) else {}
            if not name or not cls._is_temporal_field(name) or not observed_range.get("min"):
                continue
            dimensions.append(
                {
                    "field": name,
                    "kind": "temporal",
                    "min": observed_range.get("min"),
                    "max": observed_range.get("max"),
                    "basis": observed_range.get("basis") or "profile_sample",
                }
            )
        return dimensions

    @classmethod
    def _logical_dataset_profile_payload(
        cls,
        *,
        asset: KnowledgeTableAsset,
        sources: list[KnowledgeTableAsset],
        logical: dict[str, Any],
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """Aggregate source profiles without materialising the virtual union."""

        source_statuses: list[dict[str, Any]] = []
        coverage: list[dict[str, Any]] = []
        profiles_by_asset: dict[str, dict[str, Any]] = {}
        for source in sources:
            profile, state = cls._profile_source_state(source)
            source_statuses.append(state)
            if profile:
                profiles_by_asset[source.asset_id] = profile
            dimensions = cls._profile_coverage_for_source(profile)
            if dimensions:
                coverage.append(
                    {
                        "source_asset_id": source.asset_id,
                        "source_name": source.file_name,
                        "status": state["profile_status"],
                        "dimensions": dimensions,
                    }
                )

        source_count = len(sources)
        fresh_count = sum(item["profile_status"] == "fresh" for item in source_statuses)
        available_count = sum(item["profile_status"] != "missing" for item in source_statuses)
        status = "ready" if source_count and fresh_count == source_count else ("partial" if available_count else "missing")
        fields = [str(field) for field in ((logical.get("schema") or {}).get("fields") or logical.get("canonical_columns") or [])]
        columns: list[dict[str, Any]] = []
        for field in fields:
            samples: list[str] = []
            observed_ranges: list[dict[str, Any]] = []
            present_in = 0
            for profile in profiles_by_asset.values():
                for column in profile.get("columns") or []:
                    if not isinstance(column, dict) or str(column.get("name")) != field:
                        continue
                    present_in += 1
                    for value in column.get("sample_values") or []:
                        rendered = str(value)
                        if rendered not in samples and len(samples) < 5:
                            samples.append(rendered)
                    observed_range = column.get("observed_range")
                    if isinstance(observed_range, dict) and observed_range.get("min") is not None:
                        observed_ranges.append(observed_range)
            columns.append(
                {
                    "name": field,
                    "semantic_role_hint": "dimension_candidate",
                    "present_in_sources": present_in,
                    "source_count": source_count,
                    "sample_values": samples,
                    "observed_ranges": observed_ranges[:5],
                }
            )

        row_estimate = sum(int(source.rows or 0) for source in sources) if all(source.rows is not None for source in sources) else None
        now = datetime.now(BEIJING_TZ).isoformat()
        summary = {
            "status": status,
            "generated_at": now,
            "source_profiles": source_statuses,
            "coverage": coverage,
            "fresh_source_count": fresh_count,
            "profiled_source_count": available_count,
            "source_count": source_count,
            "note": "覆盖范围来自各来源的 Profile 样本；缺少或过期来源会明确标记，不以 LLM 推测补齐。",
        }
        profile_payload = {
            "asset_id": asset.asset_id,
            "kind": "logical_dataset_profile",
            "source_type": LOGICAL_CONCAT_SOURCE_TYPE,
            "file_name": asset.file_name,
            "generated_at": now,
            "status": status,
            "sampled": True,
            "shape": [row_estimate, len(fields)] if row_estimate is not None else [None, len(fields)],
            "columns": columns,
            "preview": [],
            "coverage": coverage,
            "source_freshness": source_statuses,
            "summary": summary,
        }
        return profile_payload, summary

    async def _generate_logical_dataset_profile(
        self,
        session: AsyncSession,
        asset: KnowledgeTableAsset,
        *,
        include_profile: bool,
        populate_missing_source_profiles: bool,
    ) -> dict[str, Any]:
        metadata = dict(asset.asset_metadata) if isinstance(asset.asset_metadata, dict) else {}
        logical = dict(metadata.get("logical_dataset")) if isinstance(metadata.get("logical_dataset"), dict) else {}
        source_ids = [str(item).strip() for item in logical.get("source_asset_ids") or [] if str(item).strip()]
        if len(source_ids) < 2:
            raise TableCatalogError("虚拟逻辑数据集缺少有效来源定义。")
        result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
        by_id = {item.asset_id: item for item in result.scalars()}
        sources = [by_id[source_id] for source_id in source_ids if source_id in by_id and by_id[source_id].reference_status != "removed"]
        if len(sources) != len(source_ids):
            raise TableCatalogError("虚拟逻辑数据集存在不可用来源表。")

        if populate_missing_source_profiles:
            for source in sources:
                _profile, state = self._profile_source_state(source)
                # A logical dataset summary is only reliable when every source was profiled
                # against its current file state. Existing legacy profiles without a freshness
                # marker are intentionally treated as stale and regenerated here.
                if state["profile_status"] != "fresh":
                    await self.generate_profile(session, source.asset_id, include_profile=False)
            result = await session.execute(select(KnowledgeTableAsset).where(KnowledgeTableAsset.asset_id.in_(source_ids)))
            by_id = {item.asset_id: item for item in result.scalars()}
            sources = [by_id[source_id] for source_id in source_ids]

        profile_payload, summary = self._logical_dataset_profile_payload(asset=asset, sources=sources, logical=logical)
        profile_path = self.profile_path(asset.asset_id)
        profile_path.parent.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(profile_path.write_text, json.dumps(profile_payload, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")
        logical["profile"] = summary
        logical["coverage"] = summary["coverage"]
        logical["profile_refreshed_at"] = summary["generated_at"]
        definition = {"asset_id": asset.asset_id, "name": asset.file_name, **logical}
        definition_path = Path(asset.storage_path)
        await asyncio.to_thread(definition_path.write_text, json.dumps(definition, ensure_ascii=False, indent=2), encoding="utf-8")
        asset.asset_metadata = {**metadata, "logical_dataset": logical}
        asset.profile_status = summary["status"]
        asset.profile_path = str(profile_path)
        asset.size_bytes = definition_path.stat().st_size
        asset.modified_at = datetime.now(BEIJING_TZ)
        asset.content_sha256 = await asyncio.to_thread(_sha256, definition_path)
        await session.commit()
        await session.refresh(asset)
        return self._asset_to_dict(asset, include_profile=include_profile)

    async def refresh_profiles(self, session: AsyncSession, *, limit: int = 200) -> dict[str, Any]:
        await self.ensure_catalog_populated(session, limit=limit)
        stmt = (
            select(KnowledgeTableAsset.asset_id)
            .where(KnowledgeTableAsset.knowledge_base_id == DEFAULT_KNOWLEDGE_BASE_ID)
            .order_by(KnowledgeTableAsset.updated_at.desc())
            .limit(max(1, min(limit, 1000)))
        )
        result = await session.execute(stmt)
        asset_ids = [str(asset_id) for asset_id in result.scalars()]
        generated: list[dict[str, Any]] = []
        errors: list[dict[str, str]] = []
        for asset_id in asset_ids:
            try:
                generated.append(
                    await self.generate_profile(
                        session,
                        asset_id,
                        populate_logical_source_profiles=False,
                    )
                )
            except Exception as exc:  # pragma: no cover - error payload path
                errors.append({"asset_id": asset_id, "error": str(exc)})
        return {"generated": generated, "errors": errors, "total": len(asset_ids)}

    async def register_path(
        self,
        session: AsyncSession,
        path: Path,
        *,
        virtual_path: str | None = None,
        knowledge_base_id: str = DEFAULT_KNOWLEDGE_BASE_ID,
        document_id: str | None = None,
    ) -> list[KnowledgeTableAsset]:
        await KnowledgeService(self.base_dir).ensure_default_knowledge_base(session)
        if not path.exists() or not path.is_file() or path.suffix.lower() not in TABLE_SUFFIXES:
            return []
        refs = self._dedupe_refs(self._asset_refs_for_path(
            path,
            virtual_path=virtual_path,
            document_id=document_id,
            content_sha256=_sha256(path),
        ))
        return await self._commit_asset_refs(session, refs, knowledge_base_id=knowledge_base_id)

    async def ensure_catalog_populated(self, session: AsyncSession, *, limit: int = 500) -> int:
        await KnowledgeService(self.base_dir).ensure_default_knowledge_base(session)
        count_stmt = select(KnowledgeTableAsset.asset_id).where(KnowledgeTableAsset.knowledge_base_id == DEFAULT_KNOWLEDGE_BASE_ID).limit(1)
        existing = (await session.execute(count_stmt)).scalar_one_or_none()
        if existing:
            return 0
        refs = self._dedupe_refs(self._scan_assets(limit=limit))
        await self._commit_asset_refs(session, refs, knowledge_base_id=DEFAULT_KNOWLEDGE_BASE_ID)
        return len(refs)

    async def sync_from_documents(self, session: AsyncSession, *, limit: int = 500) -> int:
        stmt = (
            select(KnowledgeDocument)
            .where(
                KnowledgeDocument.knowledge_base_id == DEFAULT_KNOWLEDGE_BASE_ID,
                KnowledgeDocument.source_type == "file_upload",
            )
            .order_by(KnowledgeDocument.updated_at.desc())
            .limit(max(1, min(limit, 2000)))
        )
        result = await session.execute(stmt)
        count = 0
        for document in result.scalars():
            path = Path(document.storage_path)
            if path.suffix.lower() not in TABLE_SUFFIXES:
                continue
            assets = await self.register_path(
                session,
                path,
                virtual_path=document.virtual_path,
                knowledge_base_id=document.knowledge_base_id,
                document_id=document.id,
            )
            count += len(assets)
        return count

    def _scan_assets(self, *, limit: int) -> list[TableAssetRef]:
        if not self.knowledge_root.exists():
            return []
        assets: list[TableAssetRef] = []
        for path in sorted(self.knowledge_root.rglob("*")):
            if len(assets) >= limit:
                break
            if not path.is_file() or path.name.startswith("."):
                continue
            suffix = path.suffix.lower()
            if suffix not in TABLE_SUFFIXES:
                continue
            try:
                relative_path = path.relative_to(self.knowledge_root)
            except ValueError:
                continue
            if not self._is_catalog_asset_path(relative_path):
                continue
            try:
                assets.extend(self._asset_refs_for_path(path, content_sha256=_sha256(path)))
            except Exception:
                continue
        return sorted(assets, key=lambda asset: (asset.modified_at, asset.file_name, asset.sheet_name or ""), reverse=True)[:limit]

    @staticmethod
    def _dedupe_refs(refs: list[TableAssetRef]) -> list[TableAssetRef]:
        deduped: dict[str, TableAssetRef] = {}
        for ref in refs:
            deduped[ref.asset_id] = ref
        return list(deduped.values())

    async def _commit_asset_refs(
        self,
        session: AsyncSession,
        refs: list[TableAssetRef],
        *,
        knowledge_base_id: str,
    ) -> list[KnowledgeTableAsset]:
        if not refs:
            return []
        last_error: IntegrityError | None = None
        for _attempt in range(3):
            assets = [
                await self._upsert_asset_ref(session, ref, knowledge_base_id=knowledge_base_id)
                for ref in refs
            ]
            try:
                await session.commit()
                refreshed: list[KnowledgeTableAsset] = []
                for asset in assets:
                    current = await session.get(KnowledgeTableAsset, asset.asset_id)
                    if current is not None:
                        refreshed.append(current)
                return refreshed
            except IntegrityError as exc:
                last_error = exc
                await session.rollback()
        raise TableCatalogError(f"Failed to update table asset catalog: {last_error}") from last_error

    @staticmethod
    def _is_catalog_asset_path(relative_path: Path) -> bool:
        parts = relative_path.parts
        if not parts:
            return False
        if parts[0] in IGNORED_TOP_LEVEL_DIRS:
            return False
        return True

    def _asset_refs_for_path(
        self,
        path: Path,
        *,
        virtual_path: str | None = None,
        document_id: str | None = None,
        content_sha256: str = "",
    ) -> list[TableAssetRef]:
        suffix = path.suffix.lower()
        stat = path.stat()
        if virtual_path is None:
            virtual_path = _virtual_path(self.knowledge_root, path)
        source_type = "excel" if suffix in {".xlsx", ".xls"} else ("tsv" if suffix == ".tsv" else "csv")
        sheets: list[str | None]
        if source_type == "excel":
            import pandas as pd

            sheets = [str(sheet) for sheet in pd.ExcelFile(path).sheet_names]
        else:
            sheets = [None]
        return [
            TableAssetRef(
                asset_id=_asset_id(virtual_path, sheet),
                path=path,
                virtual_path=virtual_path,
                file_name=path.name,
                source_type=source_type,
                sheet_name=sheet,
                size_bytes=stat.st_size,
                modified_at=_utc_datetime(stat.st_mtime),
                content_sha256=content_sha256,
                document_id=document_id,
            )
            for sheet in sheets
        ]

    async def _upsert_asset_ref(
        self,
        session: AsyncSession,
        ref: TableAssetRef,
        *,
        knowledge_base_id: str,
    ) -> KnowledgeTableAsset:
        with session.no_autoflush:
            asset = await session.get(KnowledgeTableAsset, ref.asset_id)
            if asset is None:
                stmt = select(KnowledgeTableAsset).where(
                    KnowledgeTableAsset.knowledge_base_id == knowledge_base_id,
                    KnowledgeTableAsset.virtual_path == ref.virtual_path,
                    KnowledgeTableAsset.sheet_name == ref.sheet_name,
                )
                asset = (await session.execute(stmt)).scalars().first()
        profile_path = self.profile_path(ref.asset_id)
        profile = self._read_profile(profile_path) if profile_path.exists() else None
        rows = None
        columns_count = None
        columns: list[str] = []
        if profile:
            shape = profile.get("shape") if isinstance(profile.get("shape"), list) else []
            rows = int(shape[0]) if len(shape) > 0 and shape[0] is not None else None
            columns_count = int(shape[1]) if len(shape) > 1 and shape[1] is not None else None
            columns = [str(column.get("name")) for column in profile.get("columns", []) if isinstance(column, dict) and column.get("name")]

        if asset is None:
            asset = KnowledgeTableAsset(asset_id=ref.asset_id, knowledge_base_id=knowledge_base_id)
            session.add(asset)
        asset.document_id = ref.document_id or asset.document_id
        asset.source_type = ref.source_type
        asset.file_name = ref.file_name
        asset.storage_path = str(ref.path)
        asset.virtual_path = ref.virtual_path
        asset.sheet_name = ref.sheet_name
        asset.size_bytes = ref.size_bytes
        asset.modified_at = ref.modified_at
        asset.content_sha256 = ref.content_sha256 or asset.content_sha256 or ""
        asset.profile_status = "ready" if profile else (asset.profile_status or "missing")
        asset.profile_path = str(profile_path)
        asset.rows = rows if rows is not None else asset.rows
        asset.columns_count = columns_count if columns_count is not None else asset.columns_count
        asset.columns = columns or asset.columns or []
        asset.reference_status = asset.reference_status or "pending"
        asset.asset_metadata = {
            **(asset.asset_metadata or {}),
            "catalog_source": "scan_or_import",
        }
        return asset

    def _asset_to_dict(self, asset: KnowledgeTableAsset, *, include_profile: bool) -> dict[str, Any]:
        profile = self._read_profile(Path(asset.profile_path)) if asset.profile_path and Path(asset.profile_path).exists() else None
        profile_status = str(profile.get("status") or "ready") if profile else "missing"
        result: dict[str, Any] = {
            "asset_id": asset.asset_id,
            "file_name": asset.file_name,
            "source_type": asset.source_type,
            "virtual_path": asset.virtual_path,
            "sheet_name": asset.sheet_name,
            "size_bytes": asset.size_bytes,
            "modified_at": _iso(asset.modified_at) or _iso(asset.updated_at) or "",
            "profile_status": profile_status,
            "profile_path": asset.profile_path,
            "rows": asset.rows,
            "columns_count": asset.columns_count,
            "columns": asset.columns or [],
            "reference_status": asset.reference_status,
        }
        if include_profile and profile:
            result["profile"] = profile
        logical_dataset = (asset.asset_metadata or {}).get("logical_dataset") if isinstance(asset.asset_metadata, dict) else None
        if isinstance(logical_dataset, dict):
            result["logical_dataset"] = logical_dataset
        return result

    @staticmethod
    def _read_profile(path: Path) -> dict[str, Any] | None:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None
        return payload if isinstance(payload, dict) else None

    @staticmethod
    def _ref_from_model(asset: KnowledgeTableAsset) -> TableAssetRef:
        path = Path(asset.storage_path)
        if not path.exists():
            raise TableCatalogError(f"Table asset file not found: {path}")
        return TableAssetRef(
            asset_id=asset.asset_id,
            path=path,
            virtual_path=asset.virtual_path,
            file_name=asset.file_name,
            source_type=asset.source_type,
            sheet_name=asset.sheet_name,
            size_bytes=asset.size_bytes,
            modified_at=asset.modified_at or datetime.now(timezone.utc),
            content_sha256=asset.content_sha256,
            document_id=asset.document_id,
        )

    @staticmethod
    def _derived_ref(asset_id: str, name: str, path: Path, rows: int) -> TableAssetRef:
        stat = path.stat()
        return TableAssetRef(
            asset_id=asset_id,
            path=path,
            virtual_path=f"/knowledge/.puddingclaw/derived/concat/{asset_id}/dataset.parquet",
            file_name=name,
            source_type=DERIVED_CONCAT_SOURCE_TYPE,
            sheet_name=None,
            size_bytes=stat.st_size,
            modified_at=_utc_datetime(stat.st_mtime),
            content_sha256="",
            document_id=None,
        )

    @classmethod
    def _inspect_concat_refs(cls, refs: list[TableAssetRef]) -> dict[str, Any]:
        """Return a deterministic field diff without materializing any output."""

        canonical_columns: list[str] = []
        canonical_set: set[str] = set()
        sources: list[dict[str, Any]] = []
        for index, ref in enumerate(refs):
            frame = cls._load_dataframe(ref, sample_rows=1)
            columns = [str(column).strip() for column in frame.columns]
            if not all(columns) or len(set(columns)) != len(columns):
                raise TableCatalogError(f"来源表 {ref.file_name} 存在空列名或规范化后的重复列名。")
            reserved = sorted(set(columns) & set(DERIVED_CONCAT_LINEAGE_COLUMNS))
            if reserved:
                raise TableCatalogError(f"来源表 {ref.file_name} 使用了系统保留字段: {', '.join(reserved)}")
            if index == 0:
                canonical_columns.extend(columns)
                canonical_set = set(columns)
            else:
                canonical_columns.extend(column for column in columns if column not in canonical_set)
                canonical_set.update(columns)
            sources.append({"asset_id": ref.asset_id, "file_name": ref.file_name, "sheet_name": ref.sheet_name, "columns": columns})

        baseline = set(sources[0]["columns"])
        for source in sources:
            source_columns = set(source["columns"])
            source["missing_from_baseline"] = sorted(baseline - source_columns)
            source["extra_vs_baseline"] = sorted(source_columns - baseline)
            source["missing_from_union"] = sorted(canonical_set - source_columns)

        has_schema_drift = any(source["missing_from_baseline"] or source["extra_vs_baseline"] for source in sources[1:])
        return {
            "baseline_columns": sources[0]["columns"],
            "canonical_columns": canonical_columns,
            "baseline_asset_id": refs[0].asset_id,
            "baseline_file_name": refs[0].file_name,
            "has_schema_drift": has_schema_drift,
            "sources": sources,
        }

    @classmethod
    def _concat_arrow_schema(cls, frames: list[tuple[TableAssetRef, Any]], columns: list[str]):
        """Infer one durable schema so minor monthly dtype drift cannot break Parquet writes."""

        import pandas as pd
        import pyarrow as pa

        fields: list[pa.Field] = []
        for column in columns:
            values = [frame[column] for _ref, frame in frames if column in frame]
            dtypes = [series.dtype for series in values]
            if dtypes and all(pd.api.types.is_bool_dtype(dtype) for dtype in dtypes):
                arrow_type = pa.bool_()
            elif dtypes and all(pd.api.types.is_integer_dtype(dtype) for dtype in dtypes):
                arrow_type = pa.int64()
            elif dtypes and all(pd.api.types.is_numeric_dtype(dtype) for dtype in dtypes):
                arrow_type = pa.float64()
            else:
                # Text is the only lossless common type for mixed Excel/CSV inference.
                arrow_type = pa.string()
            fields.append(pa.field(column, arrow_type, nullable=True))
        fields.extend(pa.field(column, pa.string(), nullable=False) for column in DERIVED_CONCAT_LINEAGE_COLUMNS)
        return pa.schema(fields)

    @classmethod
    def _coerce_concat_frame(cls, frame: Any, *, columns: list[str], schema: Any, ref: TableAssetRef):
        import pandas as pd

        normalized_columns = [str(column).strip() for column in frame.columns]
        frame.columns = normalized_columns
        frame = frame.reindex(columns=columns)
        for field in schema:
            if field.name not in columns:
                continue
            if str(field.type) in {"int64", "double"}:
                frame[field.name] = pd.to_numeric(frame[field.name], errors="coerce")
            elif str(field.type) == "bool":
                frame[field.name] = frame[field.name].astype("boolean")
            else:
                frame[field.name] = frame[field.name].astype("string")
        frame["_pc_source_asset_id"] = ref.asset_id
        frame["_pc_source_file_name"] = ref.file_name
        frame["_pc_source_sheet_name"] = ref.sheet_name or ""
        frame["_pc_source_virtual_path"] = ref.virtual_path
        return frame

    @classmethod
    def _materialize_concat_dataset(
        cls,
        refs: list[TableAssetRef],
        output_path: Path,
        *,
        schema_mode: str,
        canonical_columns: list[str],
    ) -> dict[str, Any]:
        """Write a vertical concat with lineage and an explicitly selected field policy."""
        import os

        import pyarrow as pa
        import pyarrow.parquet as pq

        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = output_path.with_suffix(".parquet.tmp")
        writer: pq.ParquetWriter | None = None
        total_rows = 0
        try:
            frames: list[tuple[TableAssetRef, Any]] = []
            for ref in refs:
                frame = cls._load_dataframe(ref)
                normalized_columns = [str(column).strip() for column in frame.columns]
                if not all(normalized_columns) or len(set(normalized_columns)) != len(normalized_columns):
                    raise TableCatalogError(f"来源表 {ref.file_name} 存在空列名或规范化后的重复列名。")
                current_set = set(normalized_columns)
                if schema_mode == "strict" and current_set != set(canonical_columns):
                    raise TableCatalogError("来源字段不一致，请先确认“缺失字段补空并合并”。")
                frame.columns = normalized_columns
                frames.append((ref, frame))

            if not frames:
                raise TableCatalogError("没有可合并的数据行。")
            schema = cls._concat_arrow_schema(frames, canonical_columns)
            for ref, frame in frames:
                materialized_frame = cls._coerce_concat_frame(frame, columns=canonical_columns, schema=schema, ref=ref)
                table = pa.Table.from_pandas(materialized_frame, schema=schema, preserve_index=False, safe=False)
                if writer is None:
                    writer = pq.ParquetWriter(temporary_path, schema, compression="zstd")
                writer.write_table(table)
                total_rows += len(materialized_frame)
            if writer is None:
                raise TableCatalogError("没有可合并的数据行。")
            writer.close()
            writer = None
            os.replace(temporary_path, output_path)
            return {"rows": total_rows, "canonical_columns": canonical_columns, "columns": [*canonical_columns, *DERIVED_CONCAT_LINEAGE_COLUMNS]}
        except Exception:
            if writer is not None:
                writer.close()
            temporary_path.unlink(missing_ok=True)
            raise

    def _logical_dataset_definition(
        self,
        *,
        asset_id: str,
        name: str,
        description: str,
        tags: list[str],
        routing: dict[str, Any] | None,
        sources: list[KnowledgeTableAsset],
        source_ids: list[str],
        schema_mode: str,
        output_columns: list[str],
        schema_preview: dict[str, Any],
        refreshed_at: datetime,
    ) -> dict[str, Any]:
        """Build a small, deterministic virtual-dataset contract without scanning full source data."""
        rows_known = all(source.rows is not None for source in sources)
        source_summaries = [
            {
                "asset_id": source.asset_id,
                "name": source.file_name,
                "sheet_name": source.sheet_name,
                "rows_estimate": source.rows,
                "fields": [str(column) for column in (source.columns or [])],
                "file_state": self._source_file_state(Path(source.storage_path)),
            }
            for source in sources
        ]
        profile_payload, profile_summary = self._logical_dataset_profile_payload(
            asset=KnowledgeTableAsset(asset_id=asset_id, file_name=name, source_type=LOGICAL_CONCAT_SOURCE_TYPE),
            sources=sources,
            logical={"schema": {"fields": output_columns}, "canonical_columns": output_columns},
        )
        # A definition refresh must never fabricate a fresh profile. Preserve the last aggregate
        # only when all of its source snapshots still agree with the current source files.
        profile_summary["generated_at"] = refreshed_at.isoformat()
        return {
            "formatter": "logical-data-asset",
            "version": "0.1.0",
            "kind": "vertical_union",
            "description": description.strip(),
            "tags": [str(tag).strip() for tag in tags if str(tag).strip()],
            "materialization": "virtual",
            "schema_mode": schema_mode,
            "source_asset_ids": source_ids,
            "canonical_columns": output_columns,
            "sources": source_summaries,
            "schema": {
                "fields": output_columns,
                "lineage_fields": list(DERIVED_CONCAT_LINEAGE_COLUMNS),
            },
            "coverage": profile_summary["coverage"],
            "profile": profile_summary,
            "statistics": {
                "source_count": len(sources),
                "rows_estimate": sum(int(source.rows or 0) for source in sources) if rows_known else None,
            },
            "routing": {
                "preferred_intents": (routing or {}).get("preferred_intents") or ["cross_source_aggregation", "trend", "period_comparison"],
                "direct_source_allowed": bool((routing or {}).get("direct_source_allowed", True)),
                "direct_source_when": ["user_specifies_source", "single_period_detail"],
            },
            "schema_preview": schema_preview,
            "refreshed_at": refreshed_at.isoformat(),
        }

    @classmethod
    def _read_virtual_concat(
        cls,
        refs: list[TableAssetRef],
        *,
        schema_mode: str,
        canonical_columns: list[str],
    ):
        """Resolve a virtual concat only when a consumer actually requests rows."""
        import pandas as pd

        frames: list[Any] = []
        for ref in refs:
            frame = cls._load_dataframe(ref)
            frame.columns = [str(column).strip() for column in frame.columns]
            if schema_mode == "strict" and set(frame.columns) != set(canonical_columns):
                raise TableCatalogError("虚拟逻辑数据集的来源字段已变化，请重新检查字段策略。")
            frame = frame.reindex(columns=canonical_columns)
            frame["_pc_source_asset_id"] = ref.asset_id
            frame["_pc_source_file_name"] = ref.file_name
            frame["_pc_source_sheet_name"] = ref.sheet_name or ""
            frame["_pc_source_virtual_path"] = ref.virtual_path
            frames.append(frame)
        if not frames:
            raise TableCatalogError("虚拟逻辑数据集没有可读取的来源。")
        return pd.concat(frames, ignore_index=True, copy=False)

    @staticmethod
    def _load_dataframe(asset: TableAssetRef, *, sample_rows: int | None = None):
        import pandas as pd

        nrows = max(1, sample_rows) if sample_rows else None
        if asset.source_type == DERIVED_CONCAT_SOURCE_TYPE:
            if nrows:
                import pyarrow.parquet as pq

                return pq.read_table(asset.path).slice(0, nrows).to_pandas()
            return pd.read_parquet(asset.path)
        if asset.source_type == "excel":
            return pd.read_excel(asset.path, sheet_name=asset.sheet_name, nrows=nrows)
        sep = "\t" if asset.source_type == "tsv" else ","
        try:
            return pd.read_csv(asset.path, sep=sep, nrows=nrows)
        except UnicodeDecodeError:
            return pd.read_csv(asset.path, sep=sep, encoding="gb18030", nrows=nrows)

    @staticmethod
    def _table_shape(asset: TableAssetRef, *, sample_df: Any) -> list[int] | None:
        """Return actual table shape without fully materializing large files when possible."""

        if asset.source_type == DERIVED_CONCAT_SOURCE_TYPE:
            try:
                import pyarrow.parquet as pq

                metadata = pq.ParquetFile(asset.path).metadata
                return [int(metadata.num_rows), int(metadata.num_columns)]
            except Exception:
                return [int(sample_df.shape[0]), int(sample_df.shape[1])]

        if asset.source_type == "excel" and asset.path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(asset.path, read_only=True, data_only=True)
                try:
                    sheet = workbook[asset.sheet_name] if asset.sheet_name else workbook.active
                    rows = max(int(sheet.max_row or 0) - 1, 0)
                    columns = int(sheet.max_column or sample_df.shape[1])
                    return [rows, columns]
                finally:
                    workbook.close()
            except Exception:
                return [int(sample_df.shape[0]), int(sample_df.shape[1])]

        if asset.source_type in {"csv", "tsv"}:
            try:
                with asset.path.open("rb") as handle:
                    rows = sum(1 for _ in handle)
                return [max(rows - 1, 0), int(sample_df.shape[1])]
            except Exception:
                return [int(sample_df.shape[0]), int(sample_df.shape[1])]

        return [int(sample_df.shape[0]), int(sample_df.shape[1])]

    @staticmethod
    def _column_profiles(df: Any, base_profile: dict[str, Any]) -> list[dict[str, Any]]:
        import pandas as pd

        columns: list[dict[str, Any]] = []
        dtypes = base_profile.get("dtypes", {}) if isinstance(base_profile.get("dtypes"), dict) else {}
        for column in df.columns:
            name = str(column)
            series = df[column]
            non_null = int(series.notna().sum())
            distinct_count = int(series.dropna().nunique())
            distinct_ratio = distinct_count / non_null if non_null else 0
            sample_values = [
                str(value)
                for value in series.dropna().astype(str).drop_duplicates().head(5).tolist()
            ]
            observed_range: dict[str, Any] | None = None
            non_null_series = series.dropna()
            if not non_null_series.empty:
                if pd.api.types.is_numeric_dtype(series):
                    observed_range = {
                        "min": _json_default(non_null_series.min()),
                        "max": _json_default(non_null_series.max()),
                        "basis": "profile_sample",
                    }
                elif TableAssetCatalog._is_temporal_field(name):
                    parsed = pd.to_datetime(non_null_series, errors="coerce")
                    if parsed.notna().any():
                        observed_range = {
                            "min": parsed.min().isoformat(),
                            "max": parsed.max().isoformat(),
                            "basis": "profile_sample",
                        }
            columns.append(
                {
                    "name": name,
                    "dtype": str(dtypes.get(name, series.dtype)),
                    "non_null": non_null,
                    "null_count": int(len(series) - non_null),
                    "distinct_count": distinct_count,
                    "distinct_ratio": round(distinct_ratio, 6),
                    "sample_values": sample_values,
                    "semantic_role_hint": "measure_candidate" if str(series.dtype).startswith(("int", "float")) else "dimension_candidate",
                    "observed_range": observed_range,
                }
            )
        return columns
